"""
Report Service - Generate reports in various formats
"""
from __future__ import annotations

from datetime import datetime
from typing import Dict, Any, Optional
import io
import json
import csv

import structlog
from sqlalchemy.ext.asyncio import AsyncSession

from app.schemas.analytics import AnalyticsFilter, ReportRequest
from app.services.analytics_service import AnalyticsService
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

logger = structlog.get_logger(__name__)


class ReportService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.analytics_service = AnalyticsService(db)
    
    async def generate_report(self, request: ReportRequest) -> Dict[str, Any]:
        """Generate report based on type and filters"""
        
        dashboard_stats = await self.analytics_service.get_dashboard_stats(request.filters)
        
        report_data = {
            "report_type": request.report_type,
            "generated_at": datetime.utcnow().isoformat(),
            "filters": request.filters.model_dump(),
            "kpis": dashboard_stats.kpis.model_dump(),
            "scheme_wise": dashboard_stats.scheme_wise,
            "office_wise": dashboard_stats.office_wise,
            "top_performers": dashboard_stats.top_performers,
            "low_performers": dashboard_stats.low_performers,
        }
        
        # Add specific data based on report type
        if request.report_type == "DAILY":
            report_data["trend"] = dashboard_stats.achievement_trend
        elif request.report_type == "MONTHLY":
            report_data["trend"] = dashboard_stats.achievement_trend
            report_data["recent_achievements"] = dashboard_stats.recent_achievements
        
        # Generate format
        if request.format == "JSON":
            return report_data
        elif request.format == "CSV":
            return self._build_csv(report_data)
        elif request.format == "EXCEL":
            return self._build_excel(report_data)
        elif request.format == "PDF":
            return self._build_pdf(report_data)
        else:
            return report_data
    
    async def get_dpr(self, division: str = "Nagpur City", report_date: Optional[str] = None) -> Dict[str, Any]:
        """Daily Performance Report"""
        from datetime import date
        target_date = date.fromisoformat(report_date) if report_date else date.today()
        
        filters = AnalyticsFilter(
            division=division,
            start_date=target_date,
            end_date=target_date,
        )
        
        stats = await self.analytics_service.get_dashboard_stats(filters)
        
        return {
            "report_type": "DPR",
            "date": target_date.isoformat(),
            "division": division,
            "stats": stats.model_dump(),
            "generated_at": datetime.utcnow().isoformat(),
        }
    
    async def get_monthly_consolidated(self, financial_year: str, month: int, division: str = "Nagpur City") -> Dict[str, Any]:
        """Monthly consolidated report"""
        filters = AnalyticsFilter(
            financial_year=financial_year,
            division=division,
        )
        
        stats = await self.analytics_service.get_dashboard_stats(filters)
        
        return {
            "report_type": "MONTHLY_CONSOLIDATED",
            "financial_year": financial_year,
            "month": month,
            "division": division,
            "stats": stats.model_dump(),
            "generated_at": datetime.utcnow().isoformat(),
        }

    def _build_excel(self, report_data: Dict[str, Any]) -> bytes:
        """Build a real .xlsx file from report_data. Returns raw bytes."""
        wb = Workbook()
        ws = wb.active
        ws.title = str(report_data.get("report_type", "Report"))[:31]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        ws.append(["Report Type", report_data.get("report_type", "")])
        ws.append(["Generated At", report_data.get("generated_at", "")])
        ws.append([])

        kpis = report_data.get("kpis") or {}
        if kpis:
            ws.append(["KPI", "Value"])
            for cell in ws[ws.max_row]:
                cell.font = header_font
                cell.fill = header_fill
            for key, value in kpis.items():
                ws.append([str(key), value])
            ws.append([])

        office_wise = report_data.get("office_wise") or []
        if office_wise:
            ws.append(["Office-wise Data"])
            if isinstance(office_wise, list) and office_wise and isinstance(office_wise[0], dict):
                columns = list(office_wise[0].keys())
                ws.append(columns)
                for cell in ws[ws.max_row]:
                    cell.font = header_font
                    cell.fill = header_fill
                for row in office_wise:
                    ws.append([row.get(col, "") for col in columns])

        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _build_pdf(self, report_data: Dict[str, Any]) -> bytes:
        """Build a real .pdf file from report_data. Returns raw bytes."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
        styles = getSampleStyleSheet()
        elements = []

        title = f"{report_data.get('report_type', 'Report')} Report"
        elements.append(Paragraph(title, styles["Title"]))
        elements.append(Paragraph(f"Generated At: {report_data.get('generated_at', '')}", styles["Normal"]))
        elements.append(Spacer(1, 0.5 * cm))

        kpis = report_data.get("kpis") or {}
        if kpis:
            elements.append(Paragraph("KPIs", styles["Heading2"]))
            kpi_rows = [["KPI", "Value"]] + [[str(k), str(v)] for k, v in kpis.items()]
            kpi_table = Table(kpi_rows, hAlign="LEFT")
            kpi_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]))
            elements.append(kpi_table)
            elements.append(Spacer(1, 0.5 * cm))

        office_wise = report_data.get("office_wise") or []
        if office_wise and isinstance(office_wise, list) and isinstance(office_wise[0], dict):
            elements.append(Paragraph("Office-wise Data", styles["Heading2"]))
            columns = list(office_wise[0].keys())
            table_rows = [columns] + [[str(row.get(col, "")) for col in columns] for row in office_wise]
            office_table = Table(table_rows, hAlign="LEFT")
            office_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]))
            elements.append(office_table)

        doc.build(elements)
        return buffer.getvalue()

    def _build_csv(self, report_data: Dict[str, Any]) -> bytes:
        """Build a real CSV file from report_data. Returns raw bytes (UTF-8)."""
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        writer.writerow(["Report Type", report_data.get("report_type", "")])
        writer.writerow(["Generated At", report_data.get("generated_at", "")])
        writer.writerow([])

        kpis = report_data.get("kpis") or {}
        if kpis:
            writer.writerow(["KPI", "Value"])
            for key, value in kpis.items():
                writer.writerow([key, value])
            writer.writerow([])

        office_wise = report_data.get("office_wise") or []
        if office_wise and isinstance(office_wise, list) and isinstance(office_wise[0], dict):
            writer.writerow(["Office-wise Data"])
            columns = list(office_wise[0].keys())
            writer.writerow(columns)
            for row in office_wise:
                writer.writerow([row.get(col, "") for col in columns])

        return buffer.getvalue().encode("utf-8")
