from io import BytesIO
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

TEMPLATE = (
    Path(__file__).resolve().parent.parent
    / "templates"
    / "Daily_Monitoring_Template.xlsx"
)

class ExcelExportService:

    def generate(self, reports, report_date):

        wb = load_workbook(TEMPLATE)

        # Force Excel/WPS to recalculate formulas
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True

        response_sheet = wb["Form Responses 1"]
        office_sheet = wb["Office wise"]

        if response_sheet.max_row > 1:
            response_sheet.delete_rows(2, response_sheet.max_row - 1)

        office_rows = {
            office_sheet[f"B{row}"].value: row
            for row in range(3, office_sheet.max_row + 1)
        }

        for r in reports:
            response_sheet.append([
                datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                r.office_name,
                str(report_date),
                f"{r.office_name}{report_date}",
                r.sb_opened,
                r.sb_closed,
                r.net_accounts,
                r.pli_policies,
                float(r.sum_assured),
                float(r.premium),
                r.speed_post_document,
                r.speed_post_parcel,
                r.business_post,
                r.international_letter,
                r.logistics,
                r.aadhaar_transactions,
                float(r.aadhaar_amount),
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ])

            row = office_rows.get(r.office_name)
            if row:
                    office_sheet[f"C{row}"] = str(report_date)
                    office_sheet[f"D{row}"] = f"{r.office_name}{report_date}"
                    office_sheet[f"E{row}"] = r.sb_opened
                    office_sheet[f"F{row}"] = r.sb_closed
                    office_sheet[f"G{row}"] = r.net_accounts
                    office_sheet[f"H{row}"] = r.pli_policies
                    office_sheet[f"I{row}"] = float(r.sum_assured)
                    office_sheet[f"J{row}"] = float(r.premium)
                    office_sheet[f"K{row}"] = r.speed_post_document
                    office_sheet[f"L{row}"] = r.speed_post_parcel
                    office_sheet[f"M{row}"] = r.business_post
                    office_sheet[f"N{row}"] = r.international_letter
                    office_sheet[f"O{row}"] = r.logistics
                    office_sheet[f"P{row}"] = r.aadhaar_transactions
                    office_sheet[f"Q{row}"] = float(r.aadhaar_amount)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
